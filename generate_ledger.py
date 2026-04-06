#!/usr/bin/env python3
"""
거래처 원장 자동화 스크립트
==============================
소스 데이터(Excel/CSV)를 읽어 거래처별 원장 파일을 생성합니다.

Usage:
    python3 generate_ledger.py <소스데이터파일> [출력폴더]

Example:
    python3 generate_ledger.py 거래데이터.xlsx
    python3 generate_ledger.py 거래데이터.xlsx ./output
"""

import sys
import os
import re
import shutil
from copy import copy

# openpyxl 커스텀 속성 버그 패치
import openpyxl.reader.excel as _excel_reader
_orig_read_custom = _excel_reader.ExcelReader.read_custom
def _patched_read_custom(self):
    try:
        _orig_read_custom(self)
    except TypeError:
        pass
_excel_reader.ExcelReader.read_custom = _patched_read_custom

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────────────────────
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template', '거래처원장.xlsx')
TEMPLATE_SHEET = '년.월'
DATA_ROW_START = 5    # 데이터 시작 행
DATA_ROW_END   = 28   # 데이터 끝 행 (기본 24행)
TOTAL_ROW      = 29   # 월합계 행
PAYMENT_START  = 31   # 입금 시작 행
PAYMENT_END    = 41   # 입금 끝 행
BALANCE_ROW    = 42   # 총외상매출금미납액 행
BANK_ROW       = 43   # 입금계좌 행


# ─────────────────────────────────────────────────────────────────────────────
# 컬럼명 정규화 (개행·공백 제거 후 매핑)
# ─────────────────────────────────────────────────────────────────────────────
def normalize(s):
    """컬럼명에서 공백·개행 제거"""
    return re.sub(r'[\s\n\r]+', '', str(s))

COLUMN_MAP = {
    # 정규화된 이름: 내부 키
    '거래일자':         'date',
    '매출거래처명':     'customer',
    '품목명':           'product',
    '규격':             'spec',
    '수량':             'qty',
    '단위':             'unit',
    '매출단가':         'sale_price',
    '매출공급가액':     'supply_amt',
    '매출세액':         'vat',
    '매출합계금액':     'total',
    '품목월일':         'item_date',
    '비고':             'remark',
    # 구분 열 (날짜코드 202603 형식)은 별도 처리
}


def detect_columns(df):
    """데이터프레임 컬럼을 내부 키로 매핑"""
    mapping = {}
    for col in df.columns:
        key = COLUMN_MAP.get(normalize(col))
        if key and key not in mapping:
            mapping[key] = col

    # 기간 열(YYYYMM) 탐색: 이름 있는 '구분' 또는 Unnamed 컬럼 모두 확인
    if 'period' not in mapping:
        for col in df.columns:
            sample = df[col].dropna().astype(str).str.strip()
            sample = sample[sample.str.match(r'^\d{6}$')]
            if not sample.empty:
                mapping['period'] = col
                break
    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────────────────────────────────────
def to_int(val):
    try:
        v = str(val).replace(',', '').replace(' ', '').replace('%', '')
        if v in ('', 'nan', 'NaN', 'None'):
            return 0
        return int(float(v))
    except Exception:
        return 0


def format_item_date(val):
    """'0312' → '3/12'  또는 이미 '3/12' 형식이면 그대로"""
    s = str(val).strip()
    if re.match(r'^\d{4}$', s):          # 0312
        return f"{int(s[:2])}/{int(s[2:])}"
    if re.match(r'^\d{1,2}/\d{1,2}$', s):  # 3/12
        return s
    return s


def safe_filename(name):
    """파일명에 사용 불가한 문자 제거"""
    return re.sub(r'[\\/:*?"<>|\s]+', '_', str(name)).strip('_')


def copy_row_format(ws, src_row, dst_row, max_col=16):
    """src_row의 셀 서식을 dst_row에 복사 (값은 비움)"""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font      = copy(src.font)
            dst.fill      = copy(src.fill)
            dst.border    = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
        dst.value = None


# ─────────────────────────────────────────────────────────────────────────────
# 원장 생성
# ─────────────────────────────────────────────────────────────────────────────
def create_ledger(customer, rows_df, year, month, col_map, output_dir):
    """거래처 한 곳의 원장 파일 생성"""

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[TEMPLATE_SHEET]

    n = len(rows_df)
    template_rows = DATA_ROW_END - DATA_ROW_START + 1  # 24

    # ── 행 확장 (거래 건수 > 24) ──────────────────────────────────────────
    if n > template_rows:
        extra = n - template_rows
        # 기준 서식 행(28)을 복사해 그 아래에 삽입
        ws.insert_rows(DATA_ROW_END + 1, amount=extra)
        for i in range(extra):
            copy_row_format(ws, DATA_ROW_END, DATA_ROW_END + 1 + i)

        # 행 번호 재계산
        last_data  = DATA_ROW_END + extra
        total_r    = TOTAL_ROW    + extra
        pay_start  = PAYMENT_START + extra
        pay_end    = PAYMENT_END   + extra
        balance_r  = BALANCE_ROW   + extra
        bank_r     = BANK_ROW      + extra

        # SUM 수식 갱신
        ws.cell(row=total_r, column=5).value = f'=SUM(E{DATA_ROW_START}:E{last_data})'
        ws.cell(row=total_r, column=6).value = f'=SUM(F{DATA_ROW_START}:F{last_data})'
        ws.cell(row=total_r, column=7).value = f'=SUM(G{DATA_ROW_START}:G{last_data})'

        # 미납액 수식 갱신
        pay_cells = '+'.join(f'L{r}' for r in range(pay_start, pay_end + 1))
        ws.cell(row=balance_r, column=7).value = f'=G4+G{total_r}-{pay_cells}'
    else:
        last_data = DATA_ROW_END
        total_r   = TOTAL_ROW
        pay_start = PAYMENT_START
        pay_end   = PAYMENT_END
        balance_r = BALANCE_ROW
        bank_r    = BANK_ROW

    # ── 제목 / 헤더 입력 ──────────────────────────────────────────────────
    ws['A1'].value = f"{year}년 {month}월 매출 현황({customer})"
    ws['A4'].value = '전기 외상매출금 미납액'
    ws['G4'].value = 0   # 전기 이월 잔액 (기본 0)
    ws.cell(row=total_r, column=1).value = f"{month}월합계"

    # ── 데이터 입력 ───────────────────────────────────────────────────────
    # iterrows() 대신 to_dict('records') 사용:
    #   - 루프 내 클로저가 row 변수를 레퍼런스로 캡처하는 버그 방지
    #   - 같은 거래처·같은 품목 다른 날짜도 모두 별도 행으로 기록
    records = rows_df.to_dict('records')
    for i, row_dict in enumerate(records):
        r = DATA_ROW_START + i

        def get(key, default='', _d=row_dict):   # _d로 값을 고정
            col_name = col_map.get(key)
            if col_name is None:
                return default
            val = _d.get(col_name, default)
            try:
                if pd.isna(val):
                    return default
            except (TypeError, ValueError):
                pass
            return val if val is not None else default

        ws.cell(row=r, column=1).value  = customer
        ws.cell(row=r, column=2).value  = get('product')
        ws.cell(row=r, column=3).value  = to_int(get('qty'))
        ws.cell(row=r, column=4).value  = to_int(get('sale_price'))
        ws.cell(row=r, column=5).value  = to_int(get('supply_amt'))
        ws.cell(row=r, column=6).value  = to_int(get('vat'))
        ws.cell(row=r, column=7).value  = to_int(get('total'))
        ws.cell(row=r, column=11).value = format_item_date(get('item_date'))
        ws.cell(row=r, column=14).value = get('remark')

    # 남은 빈 행 초기화 (template_rows > n 인 경우)
    for i in range(n, template_rows):
        r = DATA_ROW_START + i
        for col in (1, 2, 3, 4, 5, 6, 7, 11, 14):
            ws.cell(row=r, column=col).value = None

    # ── 저장 ──────────────────────────────────────────────────────────────
    fname = safe_filename(customer) + '.xlsx'
    out_path = os.path.join(output_dir, fname)
    wb.save(out_path)
    wb.close()
    del wb   # ← 참조 즉시 제거 → GC가 메모리 회수
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# 소스 파일 로드
# ─────────────────────────────────────────────────────────────────────────────
def load_source(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        # 헤더가 여러 줄일 수 있으므로 첫 번째 시트 전체를 읽어 자동 감지
        df = pd.read_excel(filepath, dtype=str, header=0)
    elif ext == '.csv':
        df = pd.read_csv(filepath, dtype=str, encoding='utf-8-sig')
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")
    df = df.fillna('')
    return df


def detect_year_month(df, col_map):
    """구분 컬럼(YYYYMM)에서 연도·월 추출"""
    period_col = col_map.get('period')
    if period_col:
        for val in df[period_col].dropna():
            s = str(val).strip()
            if re.match(r'^\d{6}$', s):
                return int(s[:4]), int(s[4:])
    # 거래일자에서 추출
    date_col = col_map.get('date')
    if date_col:
        for val in df[date_col].dropna():
            s = str(val).strip().replace('-', '').replace('/', '')
            if re.match(r'^\d{8}$', s):
                return int(s[:4]), int(s[4:6])
    return 2026, 3   # fallback


# ─────────────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    source_file = sys.argv[1]
    output_dir  = sys.argv[2] if len(sys.argv) > 2 else 'output'

    if not os.path.isfile(source_file):
        print(f"[오류] 파일을 찾을 수 없습니다: {source_file}")
        sys.exit(1)

    if not os.path.isfile(TEMPLATE_PATH):
        print(f"[오류] 템플릿 파일을 찾을 수 없습니다: {TEMPLATE_PATH}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    print(f"▶ 소스 파일 읽는 중: {source_file}")
    df = load_source(source_file)
    print(f"  → {len(df)}건 로드 완료  (컬럼 {len(df.columns)}개)")

    col_map = detect_columns(df)
    print(f"  → 감지된 컬럼: {col_map}")

    customer_col = col_map.get('customer')
    if not customer_col:
        print("[오류] '매출 거래처명' 컬럼을 찾을 수 없습니다. 컬럼명을 확인하세요.")
        sys.exit(1)

    year, month = detect_year_month(df, col_map)
    print(f"  → 대상 기간: {year}년 {month}월\n")

    groups = df.groupby(customer_col)
    total  = len(groups)
    ok = 0

    for idx, (customer, grp) in enumerate(groups, 1):
        customer = str(customer).strip()
        if not customer:
            continue
        try:
            out = create_ledger(customer, grp.reset_index(drop=True),
                                year, month, col_map, output_dir)
            print(f"  [{idx:3}/{total}] ✓ {out}")
            ok += 1
        except Exception as e:
            print(f"  [{idx:3}/{total}] ✗ {customer} → {e}")

    print(f"\n완료: {ok}/{total}개 파일 생성 → {os.path.abspath(output_dir)}/")


if __name__ == '__main__':
    main()
